import os, json, torch, datetime, io
from torchvision import models, transforms
from PIL import Image
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
from tqdm import tqdm

from utils.logger import get_logger

# === LOAD ENVIRONMENT CONFIG ===
load_dotenv()
logger = get_logger("predict")

MODEL_PATH = os.getenv("MODEL_PATH", "haring_resnet18_model.pth")
PREDICTION_PATH = os.getenv("PREDICTION_PATH", "prediction-images")
THRESHOLD = float(os.getenv("THRESHOLD", 0.5))
SAVE_PREDICTIONS_PATH = os.getenv("SAVE_PREDICTIONS_PATH", "prediction_results")

# Ensure output directory exists
os.makedirs(SAVE_PREDICTIONS_PATH, exist_ok=True)

# Create a timestamped output JSON filename (avoids overwriting old results)
timestamp = datetime.datetime.now().strftime("%m_%d_%Y_%H_%M_%S")
OUT_JSON = os.path.join(SAVE_PREDICTIONS_PATH, f"predictions_{timestamp}.json")
OUT_XLSX = os.path.join(SAVE_PREDICTIONS_PATH, f"predictions_{timestamp}.xlsx")

logger.info(f"Model path: {MODEL_PATH}")
logger.info(f"Prediction path: {PREDICTION_PATH}")
logger.info(f"Confidence threshold: {THRESHOLD}")
logger.info(f"Predictions will be saved to: {SAVE_PREDICTIONS_PATH}")

# === LOAD MODEL AND LABELS ===
try:
    # Load model checkpoint (weights + motif mapping)
    checkpoint = torch.load(MODEL_PATH, map_location="cpu")
    
    # Retrieve motif mappings
    motif_to_idx = checkpoint["motif_to_idx"]
    idx_to_motif = {v: k for k, v in motif_to_idx.items()}
    
    num_classes = len(motif_to_idx)
    logger.info(f"Loaded model with {num_classes} motif classes.")
except Exception as e:
    logger.error(f"Failed to load model or checkpoint: {e}")
    raise SystemExit(e)

# Rebuild model structure (must match training architecture)
model = models.resnet18()
model.fc = torch.nn.Linear(model.fc.in_features, num_classes)
model.load_state_dict(checkpoint["model_state_dict"])
model.eval()

# === IMAGE PREPROCESSING TRANSFORMS ===
# These must match the preprocessing used during training for consistency.
tfms = transforms.Compose([
    transforms.Resize((224, 224)),
    transforms.ToTensor(),
    transforms.Normalize([0.485, 0.456, 0.406],
                        [0.229, 0.224, 0.225])
])

# === PREDICTION HELPER ===
def predict_image(img_path):
    """Run prediction for a single image and return motif:confidence mapping.
    Args:
        img_path (str): Path to the image file.
    Returns:
        dict: {motif: confidence} for motifs above threshold.
    """
    try:
        img = Image.open(img_path).convert("RGB")
    except Exception as e:
        logger.error(f"Failed to open image {img_path}: {e}")
        return {}
    
    # Apply preprocessing and add batch dimension (shape [1, 3, 224, 224])
    x = tfms(img).unsqueeze(0)
    
    # Forward pass (no gradient tracking needed during inference)
    with torch.no_grad():
        logits = model(x)
        probs = torch.sigmoid(logits).squeeze().numpy()
    
    # Map probabilities to motif names
    results = {idx_to_motif[i]: float(probs[i]) for i in range(num_classes)}
    
    # Filter motifs below the confidence threshold and sort by confidence
    predicted = {k: v for k, v in results.items() if v >= THRESHOLD}
    sorted_pred = dict(sorted(predicted.items(), key=lambda x: x[1], reverse=True))

    return sorted_pred

# === MAIN PREDICTION HANDLER ===
def run_predictions(path):
    """Run predictions on a single image or all images in a folder."""
    if os.path.isdir(path):
        # If given a folder, iterate over all image files
        logger.info(f"Running predictions on folder: {path}")
        all_preds = {}
        
        # wrap file iteration in tqdm for progress bar
        image_files = [f for f in os.listdir(path) if f.lower().endswith((".jpg", ".jpeg", ".png"))]
        for fname in tqdm(image_files, desc="Predicting", unit="image"):
        # for fname in os.listdir(path):
        #     if fname.lower().endswith((".jpg", ".jpeg", ".png")):
                img_path = os.path.join(path, fname)
                all_preds[fname] = predict_image(img_path)
        logger.info(f"Completed predictions for {len(all_preds)} images.")
        return all_preds
    else:
        # If given a single image file
        logger.info(f"Running prediction on single image: {path}")
        return {os.path.basename(path): predict_image(path)}

# === EXECUTION ===
predictions = run_predictions(PREDICTION_PATH)

# Save to JSON
with open(OUT_JSON, "w") as f:
    json.dump(predictions, f, indent=2)
logger.info(f"Saved predictions to {OUT_JSON}")

# === SAVE TO EXCEL ===
wb = Workbook()
ws = wb.active
ws.title = "Predictions"

# Header row
headers = ["Object Identifier", "Image", "Predicted Motifs", "Confidence Scores"]
ws.append(headers)

# --- Set initial column widths ---
# A: Object ID | B: Image | C: Motifs | D: Scores
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 60
ws.column_dimensions["D"].width = 30

# Define thumbnail target size (pixels)
thumb_width, thumb_height = 150, 150

# Style for wrapping and centered alignment
wrap_center = Alignment(wrap_text=True, vertical="center")

# show Excel writing progress as well
for fname, preds in tqdm(predictions.items(), desc="Writing Excel", unit="row"):
# for fname, preds in predictions.items():
    object_id = os.path.splitext(fname)[0]

    if preds:
        motifs = ", ".join(preds.keys())
        scores = ", ".join([f"{v:.2f}" for v in preds.values()])
    else:
        motifs, scores = "—", "—"

    ws.append([object_id, "", motifs, scores])
    row = ws.max_row

    img_path = os.path.join(PREDICTION_PATH, fname)
    try:
        # --- Resize and embed image ---
        pil_img = Image.open(img_path)
        pil_img.thumbnail((thumb_width, thumb_height))

        img_bytes = io.BytesIO()
        pil_img.save(img_bytes, format="PNG")
        img_bytes.seek(0)

        xl_img = XLImage(img_bytes)
        ws.add_image(xl_img, f"B{row}")

        # --- Adjust row height dynamically ---
        ws.row_dimensions[row].height = thumb_height * 0.75  # convert px→pt

    except Exception as e:
        logger.warning(f"Could not embed image {fname}: {e}")

    # --- Apply wrapping + alignment to text cells ---
    ws[f"A{row}"].alignment = wrap_center
    ws[f"C{row}"].alignment = wrap_center
    ws[f"D{row}"].alignment = wrap_center

# --- Style header row for readability ---
header_font = Font(bold=True)
header_align = Alignment(horizontal="center", vertical="center")
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.alignment = header_align

# Freeze header row
ws.freeze_panes = "A2"

# Save Excel file
wb.save(OUT_XLSX)
logger.info(f"Excel file saved: {OUT_XLSX}")
print(f"Excel file saved: {OUT_XLSX}")

# --- Log summary of predictions ---
for fname, preds in predictions.items():
    if not preds:
        logger.info(f"{fname}: No motifs above threshold.")
    else:
        logger.info(f"{fname}: {len(preds)} motifs predicted.")